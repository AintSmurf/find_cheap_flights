import pandas as pd
import os
from openpyxl import load_workbook


class CovertToCSV:
    def __init__(self) -> None:
        pass

    def convert(self, ls):
        filename = "urls.csv"
        new_df = pd.DataFrame(ls)
        new_df.to_csv(filename, index=False)

    def convert_to_excel(self):
        # Read the original CSV file into a DataFrame
        df = pd.read_csv("urls.csv")

        # Write the modified DataFrame to a new xlsx file
        df.to_excel("flights.xlsx")

        # Modify the width of the columns here
        wb = load_workbook("flights.xlsx")
        ws = wb["Sheet1"]
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

        for col in ws["C"][1:]:
            col.style = "Hyperlink"
            col.hyperlink = col.value
        wb.save("flights.xlsx")
